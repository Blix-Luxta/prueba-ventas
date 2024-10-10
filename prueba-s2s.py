from openpyxl import load_workbook
import os
import sys

def procesar_excel(ruta_excel, ruta_texto):
    # Cargar el libro de Excel
    wb = load_workbook(ruta_excel, keep_vba=True)
    
    # Acceder a la hoja "Base Avaya"
    hoja_base = wb["Base Avaya"]
    
    # Determinar la última fila con datos en la base actual
    ultima_fila = hoja_base.max_row
    
    # Guardar las fórmulas originales
    formulas_originales = {}
    for col in range(30, 34):  # AD hasta AG
        celda = hoja_base.cell(row=2, column=col)
        if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
            formulas_originales[col] = celda.value
    
    # Borrar contenido desde B2 hasta AC[última_fila]
    print("Borrando contenido existente...")
    for fila in range(2, ultima_fila + 1):
        for col in range(1, 29):  # A hasta AC (columna 29)
            hoja_base.cell(row=fila, column=col).value = None
    
    # Leer el archivo de texto
    print("Leyendo archivo de texto...")
    with open(ruta_texto, 'r', encoding='cp1252') as file:
        datos = [linea.strip().split('\t') for linea in file]
    
    # Pegar los nuevos datos
    print(f"Pegando {len(datos)} filas de datos...")
    for i, fila in enumerate(datos):
        for j, valor in enumerate(fila):
            hoja_base.cell(row=i+1, column=j+1).value = valor
    
    # Extender las fórmulas (AD:AG)
    print("Extendiendo fórmulas...")
    nueva_ultima_fila = len(datos) + 1
    for col, formula_original in formulas_originales.items():
        for fila in range(2, nueva_ultima_fila + 1):
            nueva_formula = ajustar_formula(formula_original, 2, fila)
            hoja_base.cell(row=fila, column=col).value = nueva_formula
            
    hoja_base.delete_rows(2,2)
    
    # Acceder a la hoja "Operadores S2S"
    hoja_operadores = wb["Operadores S2S"]
    
    # Para tablas dinámicas, simplemente guardamos el archivo
    # Excel actualizará las tablas dinámicas automáticamente al abrirlo
    wb.save(ruta_excel)
    print("Preparando actualización de tabla dinámica...")
    
    # Verificar y ajustar fórmulas junto a la tabla dinámica
    print("Ajustando fórmulas adicionales...")
    filas_tabla = contar_filas_tabla_dinamica(hoja_operadores)
    ajustar_formulas_adicionales(hoja_operadores, filas_tabla)
    
    # Guardar cambios
    print("Guardando cambios...")
    wb.save(ruta_excel)
    print("Proceso completado exitosamente")
    print("NOTA: Para ver la tabla dinámica actualizada, necesitarás abrir el archivo en Excel.")

def ajustar_formula(formula, fila_orig, fila_nueva):
    """Ajusta la referencia de fila en una fórmula"""
    return formula.replace(str(fila_orig), str(fila_nueva))

def contar_filas_tabla_dinamica(hoja):
    """Cuenta las filas de la tabla dinámica"""
    fila = 8
    while hoja.cell(row=fila, column=2).value is not None:
        fila += 1
    return fila - 8

def ajustar_formulas_adicionales(hoja, filas_tabla):
    """Ajusta las fórmulas adyacentes para que coincidan con las filas de la tabla"""
    for col in range(6, 10):  # F a I
        celda_original = hoja.cell(row=8, column=col)
        if celda_original.value and isinstance(celda_original.value, str) and celda_original.value.startswith('='):
            formula_original = celda_original.value
            for fila in range(9, 8 + filas_tabla):
                nueva_formula = ajustar_formula(formula_original, 8, fila)
                hoja.cell(row=fila, column=col).value = nueva_formula

if __name__ == "__main__":
    # Obtener el directorio del script
    directorio = os.path.dirname(os.path.abspath(sys.argv[0]))
    print(f"directorio: {directorio}")
    
    # Construir rutas completas
    ruta_excel = os.path.join(directorio, "Reporte Operadores s2s - CSV MERA.xlsm")
    print(f"directorio excel: {ruta_excel}")
    
    ruta_texto = os.path.join(directorio, r"scripts\AM_Rep_Agente_Skill_Intervalo (tm).txt")
    print(f"directorio Texto: {ruta_texto}")
    
    try:
        procesar_excel(ruta_excel, ruta_texto)
    except Exception as e:
        print(f"Error durante la ejecución: {str(e)}")
        input("Presiona Enter para salir...")