import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os
import sys
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color

def copiar_formato(celda_origen, celda_destino):
    """Copia el formato de una celda a otra usando el método correcto de openpyxl"""
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.number_format = celda_origen.number_format
        celda_destino.protection = copy(celda_origen.protection)
        celda_destino.alignment = copy(celda_origen.alignment)

def aplicar_formato_titulo(hoja):
    """Aplica el formato especificado al título"""
    # Combinar celdas B1:G4
    hoja.merge_cells('B1:G4')
    
    # Obtener la celda combinada
    celda_titulo = hoja['B1']
    
    # Establecer el texto
    celda_titulo.value = "OPERADORES S2S - CSV MERA"
    
    # Establecer la fuente
    celda_titulo.font = Font(
        name='Calibri',
        size=22,
        bold=True,
        color='FFFFFF'  # Color blanco
    )
    
    # Establecer la alineación
    celda_titulo.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )
    
    # Establecer el color de fondo
    celda_titulo.fill = PatternFill(
        start_color='44546A',
        end_color='44546A',
        fill_type='solid'
    )

def aplicar_formato_encabezados(hoja, columnas):
    """Aplica el formato especificado a los encabezados"""
    for col, header in enumerate(columnas, 1):
        celda = hoja.cell(row=6, column=col, value=header)
        celda.font = Font(
            bold=True,
            color='FFFFFF'  # Color blanco
        )
        celda.fill = PatternFill(
            start_color='44546A',
            end_color='44546A',
            fill_type='solid'
        )
        celda.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

def operadores_mera_s2s(archivo_origen, archivo_destino):
    try:
        # Verificar si los archivos existen
        if not os.path.exists(archivo_origen):
            print("Error", f"El archivo origen '{archivo_origen}' no existe.")
            return
        if not os.path.exists(archivo_destino):
            print("Información", f"El archivo destino '{archivo_destino}' no existe. Se creará uno vacío.")
            wb = openpyxl.Workbook()
            wb.save(archivo_destino)
            
        # Cargar los libros de Excel primero
        wb_origen = openpyxl.load_workbook(archivo_origen)
        hoja_origen = wb_origen['Operadores S2S']
            
        # Leer un archivo Excel
        df_origen = pd.read_excel(archivo_origen, 
                                sheet_name='Operadores S2S', 
                                engine='openpyxl',
                                skiprows=6,
                                usecols=range(1, 9))
        
        # Establecer nombres de columnas
        columnas = ['PCRC', 'OPERADOR', 'Cod. Agente', 'Ll. ACD', 'LOGUEO', 'Q. Ventas', 'vma', 'Supervisor']
        df_origen.columns = columnas

        # Aplicar filtros
        df_filtrado = df_origen[df_origen['Ll. ACD'] != 0]
        df_filtrado = df_filtrado[df_filtrado["OPERADOR"] != "#N/D"]
        df_filtrado = df_filtrado[~df_filtrado.apply(lambda row: row.astype(str).str.contains(r'\(en blanco\)', case=False).any(), axis=1)]

        print(f"Registros filtrados: {df_filtrado.shape[0]}")

        # Cargar archivo destino
        wb_destino = openpyxl.load_workbook(archivo_destino)
            
        # Renombrar la hoja 'Envío' a 'Eliminar' si existe
        if "Envío" in wb_destino.sheetnames:
            hoja_envio = wb_destino["Envío"]
            hoja_envio.title = "Eliminar"
            wb_destino.save(archivo_destino)
            print("Hoja 'Envío' renombrada a 'Eliminar'.")
        
        # Crear nueva hoja en destino
        if "Operadores S2S" in wb_destino.sheetnames:
            del wb_destino["Operadores S2S"]
        hoja_destino = wb_destino.create_sheet("Operadores S2S")

        # Aplicar formato al título
        aplicar_formato_titulo(hoja_destino)

        # Copiar dimensiones de columnas
        for col in range(1, min(9, hoja_origen.max_column + 1)):
            col_letter = get_column_letter(col)
            if col_letter in hoja_origen.column_dimensions:
                hoja_destino.column_dimensions[col_letter].width = hoja_origen.column_dimensions[col_letter].width

        # Aplicar formato a los encabezados en la fila 6
        aplicar_formato_encabezados(hoja_destino, columnas)

        # Lista para almacenar el ancho máximo de las columnas
        max_widths = [0] * len(columnas)

        # Copiar datos con formato empezando desde la fila 7
        for i, row in enumerate(df_filtrado.values, 7):
            for j, value in enumerate(row, 1):
                celda_destino = hoja_destino.cell(row=i, column=j, value=value)
                celda_origen = hoja_origen.cell(row=i + 6, column=j + 1)
                copiar_formato(celda_origen, celda_destino)

                # Calcular el ancho máximo
                max_widths[j - 1] = max(max_widths[j - 1], len(str(value)))

        # Ajustar el ancho de las columnas en función del contenido
        for j, width in enumerate(max_widths, 1):
            col_letter = get_column_letter(j)
            hoja_destino.column_dimensions[col_letter].width = width + 5  # Añade un margen

        # Eliminar la fila 7 y ajustar todas las filas siguientes hacia arriba
        hoja_destino.delete_rows(7)

        # Renombrar la hoja
        hoja_destino.title = "Envío"
        if "Eliminar" in wb_destino.sheetnames:
            del wb_destino["Eliminar"]

        # Guardar y cerrar los archivos
        wb_destino.save(archivo_destino)
        wb_origen.close()
        wb_destino.close()
        print("Éxito", f"Operación completada. Archivo guardado en '{archivo_destino}'.")

    except Exception as e:
        print("Error", f"Ocurrió un error: {e}")
        import traceback
        print(traceback.format_exc())

def main():
    directorio = os.path.dirname(os.path.abspath(sys.argv[0]))
    archivo_origen = os.path.join(directorio, "Reporte Operadores s2s - CSV MERA.xlsm")
    archivo_destino = os.path.join(directorio, "Reporte Operadores s2s - Mera.xlsx")

    operadores_mera_s2s(archivo_origen, archivo_destino)
    
if __name__ == "__main__":
    main()